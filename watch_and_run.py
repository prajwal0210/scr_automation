import time
import shutil
from pathlib import Path
from datetime import datetime

import pandas as pd
from watchdog.events import FileSystemEventHandler
from watchdog.observers.polling import PollingObserver
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from scr_main import run_scr_automation
from raw_pipeline import run_full_pipeline
from support_session_pipeline import run_support_session_pipeline


# =========================================================
# PATHS
# =========================================================
BASE_DIR = Path(r"C:\Users\prajw\OneDrive\Documents\Projects\scr_automation")

SCR_DIR = BASE_DIR / "raw_files"
SUPPORT_DIR = BASE_DIR / "support_session_raw"

INPUT_DIR = BASE_DIR / "input"
PROCESSING_DIR = BASE_DIR / "processing"
RAN_ALREADY_DIR = BASE_DIR / "ran_already"
LOG_PATH = BASE_DIR / "logs" / "combined_log.txt"

OUTPUT_DIR = BASE_DIR / "output"
SUPPORT_OUTPUT_DIR = OUTPUT_DIR / "support_session_scrs"

TARGET_OUTPUT_FILE = "Automation data.xlsx"
PROCESSING_FILE = PROCESSING_DIR / TARGET_OUTPUT_FILE

for d in [INPUT_DIR, PROCESSING_DIR, RAN_ALREADY_DIR, LOG_PATH.parent, OUTPUT_DIR, SUPPORT_OUTPUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)


# =========================================================
# HELPERS
# =========================================================
def wait_until_stable(file_path: Path, checks=3, delay=2):
    last_size = -1
    stable_count = 0

    for _ in range(20):
        if not file_path.exists():
            return False

        size = file_path.stat().st_size
        if size > 0 and size == last_size:
            stable_count += 1
            if stable_count >= checks:
                return True
        else:
            stable_count = 0

        last_size = size
        time.sleep(delay)

    return False


# =========================================================
# SUMMARY HELPERS
# =========================================================
def safe_str(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def format_output_date(value):
    dt = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(dt):
        return ""
    return dt.strftime("%d.%m.%Y")


def sanitize_filename(name):
    name = safe_str(name)
    invalid_chars = r'<>:"/\|?*'
    for ch in invalid_chars:
        name = name.replace(ch, "")
    return " ".join(name.split()).strip()


def build_expected_output_filename(row):
    student_given = safe_str(row.get("Given", row.get("Student Given Name", "")))
    student_surname = safe_str(row.get("Surname", row.get("Student Surname", "")))
    student_name = f"{student_given} {student_surname}".strip()

    trainer_name = safe_str(
        row.get("Trainer Name", "")
        or f'{safe_str(row.get("Trainer Given Name", ""))} {safe_str(row.get("Trainer Surname", ""))}'.strip()
    )

    raw_date = (
        row.get("Date", "")
        if "Date" in row.index else
        row.get("Date of Contact", "")
        if "Date of Contact" in row.index else
        row.get("Event Date", "")
        if "Event Date" in row.index else
        ""
    )

    date_of_contact_file = format_output_date(raw_date)

    filename = (
        f"{student_name}_"
        f"{date_of_contact_file}_"
        f"CPC30220 - Victoria V3.1-1-2 - "
        f"{trainer_name}.docx"
    )

    trainer_folder = sanitize_filename(trainer_name) or "Unknown Trainer"
    return sanitize_filename(filename), trainer_folder


def get_actual_generated_files(is_support=False):
    base_dir = SUPPORT_OUTPUT_DIR if is_support else OUTPUT_DIR

    actual_files = set()
    if not base_dir.exists():
        return actual_files

    for file_path in base_dir.rglob("*.docx"):
        actual_files.add(file_path.name)

    return actual_files


def autosize_worksheet(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter

        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_length:
                max_length = len(val)

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def style_sheet(ws):
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = center

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    autosize_worksheet(ws)


def create_scr_summary_report(is_support=False):
    """
    Creates/updates:
    1. Summary sheet
    2. Not Generated SCR sheet

    Source file:
    - normal SCR: raw_files/events_matched_with_units.xlsx
    - support SCR: support_session_raw/support_session_events_matched_with_units.xlsx
    """
    if is_support:
        source_file = SUPPORT_DIR / "support_session_events_matched_with_units.xlsx"
    else:
        source_file = SCR_DIR / "events_matched_with_units.xlsx"

    if not source_file.exists():
        print(f"Summary source file not found: {source_file}")
        return

    df = pd.read_excel(source_file)

    if df.empty:
        print(f"Summary source file is empty: {source_file}")
        return

    actual_generated_files = get_actual_generated_files(is_support=is_support)

    expected_filenames = []
    trainer_folders = []

    for _, row in df.iterrows():
        filename, trainer_folder = build_expected_output_filename(row)
        expected_filenames.append(filename)
        trainer_folders.append(trainer_folder)

    df["Expected Output Filename"] = expected_filenames
    df["Trainer Folder"] = trainer_folders
    df["SCR Generated"] = df["Expected Output Filename"].isin(actual_generated_files)

    not_generated_df = df[~df["SCR Generated"]].copy()

    # -----------------------------
    # Trainer column detection
    # -----------------------------
    trainer_col = None
    for col in ["Trainer Name", "Trainer", "Trainer Full Name"]:
        if col in df.columns:
            trainer_col = col
            break

    if trainer_col is None:
        if "Trainer Given Name" in df.columns or "Trainer Surname" in df.columns:
            df["Trainer Name Derived"] = (
                df.get("Trainer Given Name", pd.Series("", index=df.index)).fillna("").astype(str).str.strip()
                + " " +
                df.get("Trainer Surname", pd.Series("", index=df.index)).fillna("").astype(str).str.strip()
            ).str.strip()
            trainer_col = "Trainer Name Derived"
        else:
            df["Trainer Name Derived"] = df["Trainer Folder"]
            trainer_col = "Trainer Name Derived"

    # -----------------------------
    # Supervisor detection
    # -----------------------------
    supervisor_given_col = None
    for col in ["Employer contacts -> Client -> Given name", "Supervisor Given Name"]:
        if col in df.columns:
            supervisor_given_col = col
            break

    supervisor_surname_col = None
    for col in ["Employer contacts -> Client -> Surname", "Supervisor Surname"]:
        if col in df.columns:
            supervisor_surname_col = col
            break

    def build_supervisor_name(row):
        given = safe_str(row.get(supervisor_given_col, "")) if supervisor_given_col else ""
        surname = safe_str(row.get(supervisor_surname_col, "")) if supervisor_surname_col else ""
        return f"{given} {surname}".strip()

    if supervisor_given_col or supervisor_surname_col:
        df["Supervisor Name Derived"] = df.apply(build_supervisor_name, axis=1)
    else:
        df["Supervisor Name Derived"] = ""

    # -----------------------------
    # Student name helper
    # -----------------------------
    def build_student_name(row):
        given = safe_str(row.get("Given", row.get("Student Given Name", "")))
        surname = safe_str(row.get("Surname", row.get("Student Surname", "")))
        return f"{given} {surname}".strip()

    df["Student Name Derived"] = df.apply(build_student_name, axis=1)

    # -----------------------------
    # Build summary from generated SCR only
    # -----------------------------
    summary_rows = []

    generated_df = df[df["SCR Generated"]].copy()

    for trainer_name, grp in generated_df.groupby(trainer_col, dropna=False):
        trainer_name = safe_str(trainer_name)

        no_supervisor_students = grp.loc[
            grp["Supervisor Name Derived"].eq(""),
            "Student Name Derived"
        ].dropna().astype(str).str.strip().tolist()

        no_supervisor_students = [x for x in no_supervisor_students if x]
        no_supervisor_text = "\n".join(no_supervisor_students)

        summary_rows.append({
            "Trainer Name": trainer_name,
            "Counts of SCR": int(len(grp)),
            "No Supervisor": no_supervisor_text
        })

    summary_df = pd.DataFrame(summary_rows)

    if not summary_df.empty:
        summary_df = summary_df.sort_values(by=["Trainer Name"], ascending=True).reset_index(drop=True)
        total_row = pd.DataFrame([{
            "Trainer Name": "Total",
            "Counts of SCR": int(summary_df["Counts of SCR"].sum()),
            "No Supervisor": ""
        }])
        summary_df = pd.concat([summary_df, total_row], ignore_index=True)
    else:
        summary_df = pd.DataFrame([{
            "Trainer Name": "Total",
            "Counts of SCR": 0,
            "No Supervisor": ""
        }])

    # -----------------------------
    # Write sheets back to source workbook
    # -----------------------------
    wb = load_workbook(source_file)

    for sheet_name in ["Summary", "Not Generated SCR"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws_summary = wb.create_sheet("Summary")
    ws_missing = wb.create_sheet("Not Generated SCR")

    # Summary sheet
    for c_idx, col_name in enumerate(summary_df.columns, start=1):
        ws_summary.cell(row=1, column=c_idx, value=col_name)

    for r_idx, (_, row) in enumerate(summary_df.iterrows(), start=2):
        for c_idx, col_name in enumerate(summary_df.columns, start=1):
            ws_summary.cell(row=r_idx, column=c_idx, value=row[col_name])

    # Missing sheet
    missing_output_df = not_generated_df.copy()

    for c_idx, col_name in enumerate(missing_output_df.columns, start=1):
        ws_missing.cell(row=1, column=c_idx, value=col_name)

    for r_idx, (_, row) in enumerate(missing_output_df.iterrows(), start=2):
        for c_idx, col_name in enumerate(missing_output_df.columns, start=1):
            value = row[col_name]
            if pd.isna(value):
                value = ""
            ws_missing.cell(row=r_idx, column=c_idx, value=value)

    style_sheet(ws_summary)
    style_sheet(ws_missing)

    # Wrap No Supervisor column
    for cell in ws_summary["C"]:
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    autosize_worksheet(ws_summary)
    autosize_worksheet(ws_missing)

    wb.save(source_file)
    print(f"Summary sheets created in: {source_file}")


# =========================================================
# FILE DETECTION
# =========================================================
def get_scr_files():
    employer = SCR_DIR / "Employer contact (primary).csv"
    trainer = next(SCR_DIR.glob("*trainer_events*.csv"), None)
    unit = SCR_DIR / "Unit Result Export Report.csv"

    if employer.exists() and trainer and unit.exists():
        return employer, trainer, unit
    return None


def get_support_files():
    employer = SUPPORT_DIR / "Employer contact (primary).csv"
    trainer = next(SUPPORT_DIR.glob("*support_session*.xlsx"), None)
    unit = SUPPORT_DIR / "Unit Result Export Report.csv"

    if employer.exists() and trainer and unit.exists():
        return employer, trainer, unit
    return None


# =========================================================
# HANDLER
# =========================================================
class Handler(FileSystemEventHandler):
    is_running = False
    last_run_time = 0
    cooldown = 10

    def trigger(self):
        now = time.time()

        if self.is_running or (now - self.last_run_time < self.cooldown):
            return

        # =========================
        # SUPPORT SESSION FIRST
        # =========================
        support_files = get_support_files()
        if support_files:
            self.run_support(*support_files)
            return

        # =========================
        # NORMAL SCR
        # =========================
        scr_files = get_scr_files()
        if scr_files:
            self.run_scr(*scr_files)
            return

    # =========================
    # SUPPORT SESSION FLOW
    # =========================
    def run_support(self, employer, trainer, unit):
        print("Running SUPPORT SESSION pipeline...")
        self.is_running = True
        self.last_run_time = time.time()

        try:
            run_support_session_pipeline(
                employer_contact_file=employer,
                support_session_file=trainer,
                unit_result_file=unit,
            )

            self.post_process(is_support=True)

            self.archive_files([employer, trainer, unit], prefix="support_session")

        except Exception as e:
            print(f"Support session error: {e}")

        finally:
            self.is_running = False

    # =========================
    # NORMAL SCR FLOW
    # =========================
    def run_scr(self, employer, trainer, unit):
        print("Running NORMAL SCR pipeline...")
        self.is_running = True
        self.last_run_time = time.time()

        try:
            run_full_pipeline(
                employer_contact_file=employer,
                trainer_events_file=trainer,
                unit_result_file=unit,
            )

            self.post_process(is_support=False)

            self.archive_files([employer, trainer, unit], prefix="scr")

        except Exception as e:
            print(f"SCR error: {e}")

        finally:
            self.is_running = False

    # =========================
    # POST PROCESS
    # =========================
    def post_process(self, is_support=False):
        generated_file = INPUT_DIR / TARGET_OUTPUT_FILE

        if not generated_file.exists():
            raise FileNotFoundError("Automation data.xlsx not created")

        if not wait_until_stable(generated_file):
            raise RuntimeError("File not stable")

        df = pd.read_excel(generated_file)
        count = len(df)

        shutil.copy2(generated_file, PROCESSING_FILE)

        start = time.time()
        run_scr_automation(
            excel_path=PROCESSING_FILE,
            is_support_session=is_support
        )
        duration = round(time.time() - start, 2)

        # NEW: create summary + missing SCR sheets
        create_scr_summary_report(is_support=is_support)

        log = f"[{datetime.now()}] {count} records | {duration}s | {'SUPPORT' if is_support else 'SCR'}"
        print(log)

        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(log + "\n")

        # cleanup
        generated_file.unlink(missing_ok=True)
        PROCESSING_FILE.unlink(missing_ok=True)

    # =========================
    # ARCHIVE
    # =========================
    def archive_files(self, files, prefix):
        timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
        folder = RAN_ALREADY_DIR / f"{prefix}_{timestamp}"
        folder.mkdir(parents=True, exist_ok=True)

        for f in files:
            shutil.move(str(f), str(folder / f.name))

    # =========================
    # EVENTS
    # =========================
    def on_created(self, event):
        if not event.is_directory:
            self.trigger()

    def on_modified(self, event):
        if not event.is_directory:
            self.trigger()


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    observer = PollingObserver(timeout=2)

    observer.schedule(Handler(), str(SCR_DIR), recursive=False)
    observer.schedule(Handler(), str(SUPPORT_DIR), recursive=False)

    observer.start()

    print("Watching both folders:")
    print(f"- SCR: {SCR_DIR}")
    print(f"- SUPPORT: {SUPPORT_DIR}")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()

    observer.join()
