import os
import re
from copy import copy
from datetime import datetime, time
from difflib import SequenceMatcher
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


# =========================================================
# PROJECT PATHS
# =========================================================
BASE_DIR = Path(r"C:\Users\prajw\OneDrive\Documents\Projects\scr_automation")
RAW_DIR = BASE_DIR / "raw_files"
INPUT_DIR = BASE_DIR / "input"
TEMPLATES_DIR = BASE_DIR / "templates"

TEMPLATE_PATH = TEMPLATES_DIR / "Automation data.xlsx"
FINAL_OUTPUT_PATH = INPUT_DIR / "Automation data.xlsx"


def run_full_pipeline(
    employer_contact_file,
    trainer_events_file,
    unit_result_file,
    output_folder=RAW_DIR,
    template_path=TEMPLATE_PATH,
    final_output_path=FINAL_OUTPUT_PATH,
):
    """
    Full pipeline using only 3 input files:
    1. Employer contact (primary).csv
    2. trainer_events_report_YYYY-MM-DD.csv
    3. Unit Result Export Report 1.csv

    Output:
    - Automation data.xlsx -> saved into INPUT_DIR
    """

    output_folder = Path(output_folder)
    template_path = Path(template_path)
    final_output_path = Path(final_output_path)

    output_folder.mkdir(parents=True, exist_ok=True)
    final_output_path.parent.mkdir(parents=True, exist_ok=True)

    # =========================================================
    # OUTPUT FILE PATHS (OPTIONAL DEBUG FILES)
    # =========================================================
    unit_result_output_file = output_folder / "unit_result_with_employee_contacts.xlsx"
    cleaned_events_output_file = output_folder / "cleaned_event_titles.xlsx"
    matched_output_file = output_folder / "events_matched_with_units.xlsx"

    # =========================================================
    # HELPER FUNCTIONS
    # =========================================================
    def clean_columns(df):
        df.columns = (
            df.columns.astype(str)
            .str.strip()
            .str.replace("\n", "", regex=False)
            .str.replace("\r", "", regex=False)
        )
        return df

    def make_primary_key(value):
        if pd.isna(value):
            return ""

        value = str(value).strip()
        value = re.sub(r"\s+", " ", value)
        value = re.sub(r"[^A-Za-z0-9 ]", "", value)
        value = value.replace(" ", "")
        value = value.lower()
        return value

    def create_unique_key(text):
        if pd.isna(text):
            return ""

        text = str(text).strip()
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"[^A-Za-z0-9 ]", "", text)
        text = text.replace(" ", "")
        text = text.lower()
        return text

    def join_unique_unit_codes(series):
        values = []
        seen = set()

        for val in series:
            if pd.isna(val):
                continue
            val = str(val).strip()
            if val and val not in seen:
                seen.add(val)
                values.append(val)

        return ", ".join(values)

    def clean_event_title(text):
        if pd.isna(text):
            return ""

        text = str(text).strip()
        text = re.sub(r"\s+", " ", text).strip()

        pattern = r"""
            \s*
            [\-/|:,;]*\s*
            [\(\[\{]?\s*
            site\s*[-/]?\s*visit
            \s*[\)\]\}]?
        """

        text = re.sub(pattern, "", text, flags=re.IGNORECASE | re.VERBOSE)
        text = re.sub(r"^[\s\-/|:,;]+", "", text)
        text = re.sub(r"[\s\-/|:,;]+$", "", text)
        text = re.sub(r"\s+", " ", text).strip()

        return text

    def similarity_score(a, b):
        return SequenceMatcher(None, str(a), str(b)).ratio() * 100

    def get_best_fuzzy_match(left_key, right_keys, threshold=90):
        if not left_key:
            return None, None

        best_match = None
        best_score = 0

        for rk in right_keys:
            score = similarity_score(left_key, rk)
            if score > best_score:
                best_score = score
                best_match = rk

        if best_score >= threshold:
            return best_match, round(best_score, 2)

        return None, round(best_score, 2)

    def split_trainer_name(name):
        if pd.isna(name):
            return "", ""

        parts = str(name).strip().split(" ", 1)
        given = parts[0]
        surname = parts[1] if len(parts) > 1 else ""
        return given, surname

    def parse_time_value(value):
        if pd.isna(value) or value is None:
            return None

        value = str(value).strip()
        if value == "":
            return None

        value = re.sub(r"(?i)\s*([ap]m)$", lambda m: " " + m.group(1).upper(), value)
        value = re.sub(r"\s+", " ", value).strip()

        possible_formats = [
            "%I:%M %p",   # 07:00 AM
            "%I:%M%p",    # 07:00AM
            "%H:%M",      # 07:00
            "%I.%M %p",   # 07.00 AM
            "%I.%M%p",    # 07.00AM
        ]

        for fmt in possible_formats:
            try:
                dt = datetime.strptime(value, fmt)
                return time(dt.hour, dt.minute)
            except Exception:
                pass

        return None

    # =========================================================
    # STEP 1: PROCESS UNIT RESULT + EMPLOYER CONTACT
    # =========================================================
    print("Step 1 started: Processing Unit Result and Employer Contact...")

    left_df = pd.read_csv(unit_result_file, skiprows=1, encoding="utf-8", engine="python")
    right_df = pd.read_csv(employer_contact_file, encoding="cp1252", engine="python")

    left_df = clean_columns(left_df)
    right_df = clean_columns(right_df)

    left_df = left_df[
        left_df["Contract/Enrolment Status"].astype(str).str.strip().eq("Active")
    ].copy()

    left_df["Primary key"] = left_df["Employer"].apply(make_primary_key)
    right_df["Primary key"] = right_df["Name"].apply(make_primary_key)

    right_map = right_df[
        [
            "Primary key",
            "Name",
            "Employer contacts -> Client -> Surname",
            "Employer contacts -> Client -> Given name",
        ]
    ].copy()

    right_map = right_map.drop_duplicates(subset=["Primary key"], keep="first")

    merged_df = left_df.merge(right_map, on="Primary key", how="left")

    final_columns = [
        "Surname",
        "Given",
        "Other",
        "Code",
        "Employer",
        "Unit Code",
        "Primary key",
        "Name",
        "Employer contacts -> Client -> Surname",
        "Employer contacts -> Client -> Given name",
    ]

    final_columns_existing = [col for col in final_columns if col in merged_df.columns]
    final_df = merged_df[final_columns_existing].copy()

    group_cols = [
        "Surname",
        "Given",
        "Other",
        "Code",
        "Employer",
        "Primary key",
        "Name",
        "Employer contacts -> Client -> Surname",
        "Employer contacts -> Client -> Given name",
    ]

    group_cols_existing = [col for col in group_cols if col in final_df.columns]

    collapsed_df = (
        final_df.groupby(group_cols_existing, dropna=False, as_index=False)["Unit Code"]
        .agg(join_unique_unit_codes)
    )

    final_output_columns = [
        "Surname",
        "Given",
        "Other",
        "Code",
        "Employer",
        "Unit Code",
        "Primary key",
        "Name",
        "Employer contacts -> Client -> Surname",
        "Employer contacts -> Client -> Given name",
    ]

    final_output_columns = [col for col in final_output_columns if col in collapsed_df.columns]
    collapsed_df = collapsed_df[final_output_columns]

    # Optional debug output
    # with pd.ExcelWriter(unit_result_output_file, engine="openpyxl") as writer:
    #     collapsed_df.to_excel(writer, sheet_name="Mapped Contacts", index=False)

    print(f"Step 1 completed. Output created: {unit_result_output_file}")

    # =========================================================
    # STEP 2: PROCESS TRAINER EVENTS
    # =========================================================
    print("Step 2 started: Cleaning trainer event titles...")

    events_df = pd.read_csv(trainer_events_file, encoding="utf-8", engine="python")
    events_df = clean_columns(events_df)

    events_df["Cleaned Event Title"] = events_df["Event Title"].apply(clean_event_title)

    def has_site_visit(text):
        if pd.isna(text):
            return False
        text = str(text).lower()
        return "site visit" in text or "sitevisit" in text

    events_df["has_site_visit"] = events_df["Event Title"].apply(has_site_visit)
    events_df = events_df.sort_values(by="has_site_visit", ascending=False)
    events_df = events_df.drop_duplicates(subset=["Cleaned Event Title"], keep="first")
    events_df.drop(columns=["has_site_visit"], inplace=True)

    events_df["Unique key"] = events_df["Cleaned Event Title"].apply(create_unique_key)

    # Optional debug output
    # with pd.ExcelWriter(cleaned_events_output_file, engine="openpyxl") as writer:
    #     events_df.to_excel(writer, sheet_name="Cleaned Events", index=False)

    print(f"Step 2 completed. Output created: {cleaned_events_output_file}")

    # =========================================================
    # STEP 3: FINAL MATCH
    # =========================================================
    print("Step 3 started: Matching cleaned events with units result...")

    left_match_df = clean_columns(events_df.copy())
    right_match_df = clean_columns(collapsed_df.copy())

    if "Cleaned Event Title" not in left_match_df.columns:
        raise ValueError("Column 'Cleaned Event Title' not found in cleaned events file.")

    left_match_df["Unique key"] = left_match_df["Cleaned Event Title"].apply(create_unique_key)

    required_right_cols = ["Given", "Surname"]
    missing_cols = [col for col in required_right_cols if col not in right_match_df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns in right table: {missing_cols}")

    right_match_df["Unique key"] = (
        right_match_df["Given"].fillna("").astype(str).str.strip()
        + right_match_df["Surname"].fillna("").astype(str).str.strip()
    ).apply(create_unique_key)

    final_merged_df = left_match_df.merge(
        right_match_df,
        on="Unique key",
        how="left",
        suffixes=("", "_units"),
    )

    right_keys = right_match_df["Unique key"].dropna().astype(str).unique().tolist()
    right_lookup = right_match_df.drop_duplicates(subset=["Unique key"]).set_index("Unique key")

    final_merged_df["Match Type"] = (
        final_merged_df["Name"].apply(lambda x: "Exact Match" if pd.notna(x) else "Unmatched")
        if "Name" in final_merged_df.columns
        else "Unmatched"
    )
    final_merged_df["Fuzzy Score"] = None
    final_merged_df["Matched Unique key"] = final_merged_df["Unique key"]

    for idx, row in final_merged_df[final_merged_df["Match Type"] == "Unmatched"].iterrows():
        left_key = row["Unique key"]
        best_key, best_score = get_best_fuzzy_match(left_key, right_keys, threshold=90)

        if best_key is not None and best_key in right_lookup.index:
            matched_row = right_lookup.loc[best_key]

            for col in right_match_df.columns:
                if col == "Unique key":
                    continue
                final_merged_df.at[idx, col] = matched_row[col]

            final_merged_df.at[idx, "Match Type"] = "Fuzzy Match"
            final_merged_df.at[idx, "Fuzzy Score"] = best_score
            final_merged_df.at[idx, "Matched Unique key"] = best_key

    with pd.ExcelWriter(matched_output_file, engine="openpyxl") as writer:
        final_merged_df.to_excel(writer, sheet_name="Matched Output", index=False)

    print(f"Step 3 completed. Final output created: {matched_output_file}")
    print(f"Left rows: {len(left_match_df)}")
    print(f"Final merged rows: {len(final_merged_df)}")
    print(final_merged_df["Match Type"].value_counts(dropna=False))

    # =========================================================
    # STEP 4: FORMAT + LOAD INTO TEMPLATE
    # =========================================================
    print("Step 4 started: Preparing final output and populating template...")

    if "Trainer Name" in final_merged_df.columns:
        final_merged_df[["Trainer Given Name", "Trainer Surname"]] = (
            final_merged_df["Trainer Name"].apply(lambda x: pd.Series(split_trainer_name(x)))
        )

    rename_mapping = {
        "Surname": "Student Surname",
        "Given": "Student Given Name",
        "Employer": "Employer Company",
        "Employer contacts -> Client -> Surname": "Supervisor Surname",
        "Employer contacts -> Client -> Given name": "Supervisor Given Name",
        "Date": "Date of Contact",
        "Time": "Start Time",
        "Unit Code": "Commenced",
    }

    final_df = final_merged_df.rename(columns=rename_mapping).copy()

    for col in ["Location", "End Time"]:
        if col not in final_df.columns:
            final_df[col] = ""

    if "Student Surname" in final_df.columns and "Student Given Name" in final_df.columns:
        final_df = final_df[
            ~(
                final_df["Student Surname"].fillna("").astype(str).str.strip().eq("")
                & final_df["Student Given Name"].fillna("").astype(str).str.strip().eq("")
            )
        ].copy()

    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb["Sheet1"]

    # -------------------------
    # HEADER MAP
    # -------------------------
    header_row = 1
    template_headers = {}

    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if header_value is not None:
            template_headers[str(header_value).strip()] = col_idx

    # -------------------------
    # TEMPLATE DATA ROW STYLE
    # -------------------------
    start_row = header_row + 1
    style_row = start_row

    rows_needed = len(final_df)
    existing_template_data_rows = ws.max_row - start_row + 1

    if rows_needed > existing_template_data_rows:
        ws.insert_rows(ws.max_row + 1, amount=rows_needed - existing_template_data_rows)

    # -------------------------
    # CLEAR OLD DATA AREA
    # -------------------------
    for row in range(start_row, start_row + max(existing_template_data_rows, rows_needed)):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col_idx).value = None

    # -------------------------
    # COPY STYLE ROW
    # -------------------------
    for target_row in range(start_row, start_row + rows_needed):
        if style_row in ws.row_dimensions:
            ws.row_dimensions[target_row].height = ws.row_dimensions[style_row].height

        for col_idx in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=style_row, column=col_idx)
            target_cell = ws.cell(row=target_row, column=col_idx)

            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.number_format:
                target_cell.number_format = copy(source_cell.number_format)

            if source_cell.font:
                target_cell.font = copy(source_cell.font)

            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)

            if source_cell.border:
                target_cell.border = copy(source_cell.border)

            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)

    # -------------------------
    # WRITE DATA
    # -------------------------
    for i, (_, row_data) in enumerate(final_df.iterrows()):
        excel_row = start_row + i

        for col_name, col_idx in template_headers.items():
            if col_name in final_df.columns:
                ws.cell(row=excel_row, column=col_idx).value = row_data[col_name]

    # -------------------------
    # REMOVE EMPTY STUDENTS
    # -------------------------
    surname_col = template_headers.get("Student Surname")
    given_col = template_headers.get("Student Given Name")

    if surname_col and given_col:
        rows_to_delete = []

        for row_num in range(ws.max_row, start_row - 1, -1):
            surname_val = ws.cell(row=row_num, column=surname_col).value
            given_val = ws.cell(row=row_num, column=given_col).value

            surname_blank = surname_val is None or str(surname_val).strip() == ""
            given_blank = given_val is None or str(given_val).strip() == ""

            if surname_blank and given_blank:
                rows_to_delete.append(row_num)

        for row_num in rows_to_delete:
            ws.delete_rows(row_num, 1)

    # -------------------------
    # FORMAT TIME COLUMNS
    # -------------------------
    time_columns = ["Start Time", "End Time"]

    for col_name in time_columns:
        if col_name in template_headers:
            col_idx = template_headers[col_name]

            for row_num in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                parsed_time = parse_time_value(cell.value)

                if parsed_time is not None:
                    cell.value = parsed_time
                    cell.number_format = "hh:mm AM/PM"

    # -------------------------
    # SAVE FINAL OUTPUT TO INPUT FOLDER
    # -------------------------
    wb.save(final_output_path)

    print("Step 4 completed.")
    print(f"Final file saved: {final_output_path}")

    return final_df


if __name__ == "__main__":
    run_full_pipeline(
        employer_contact_file=RAW_DIR / "Employer contact (primary).csv",
        trainer_events_file=RAW_DIR / "trainer_events_report_2026-03-23.csv",
        unit_result_file=RAW_DIR / "Unit Result Export Report 1.csv",
    )